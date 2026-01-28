import pandas as pd
import config

from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule

from seleccionar_ruta import seleccionar_ruta
from tratarHoras import tratarHoras
from tratar_certificaciones import tratar_certificaciones
from tratarHorasCompleto import tratarHorasCompleto
from tratarHorasEscaleras import tratarHorasEscaleras


def tratarHorasCompleto1(fichero, horasMesAsc,horasEscaleras,horasMesEsc, certificacion,personal,opcion_seleccionada):

  variable1=1
  ascensores=tratarHorasCompleto(fichero, horasMesAsc,certificacion,personal,opcion_seleccionada)
  variable1=2
  escaleras=tratarHorasEscaleras(HorasEscaleras,HorasMesEsc,certificacion,personal,opcion_seleccionada)

   ruta_salida = seleccionar_ruta()
    if not ruta_salida:
        print("No se seleccionó ruta de salida")
        exit()
    ascensores.to_excel(ruta_salida, sheet_name='hoja1', index=False)
    escaleras.to_excel(ruta_salida, sheet_name='hoja2', index=False)
    
    # Abrir con openpyxl y aplicar formato
    wb = load_workbook(ruta_salida)
    ws = wb.active
    
    # Buscar la columna de estado_certificacion
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "estado_certificacion":
            col_idx = col[0].column_letter
            break
    
    # Regla: si contiene "sin", poner en rojo
    formula = f'ISNUMBER(SEARCH("sin", ${col_idx}2))'
    red_font = Font(color="9C0006")
    rule = FormulaRule(formula=[formula], font=red_font)
    
    # Aplicar formato desde fila 2 hasta el final
    ws.conditional_formatting.add(f"{col_idx}2:{col_idx}{ws.max_row}", rule)
    
    # Guardar archivo final
    wb.save(ruta_salida)
    
    print("Archivo guardado correctamente")
    
