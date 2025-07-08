import pandas as pd
import config

from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule

from seleccionar_ruta import seleccionar_ruta
from tratarHoras import tratarHoras
from tratar_certificaciones import tratar_certificaciones


def tratarHorasCompleto (fichero, fichero1,fichero2,fichero3,opcion_seleccionada):

    horasTratadas=tratarHoras(fichero,opcion_seleccionada)
    # ---------- 1. Cargar resumen 12 meses ----------
    resumen = horasTratadas

    # ---------- 2. Cargar horas del mes actual ----------
    mes_actual = pd.read_excel(fichero1)
    '''mes_actual["Activity"] = mes_actual["Activity"].str.strip().str.lower()'''
    mes_actual.rename(columns={"Nº Personal":"id"}, inplace=True)
    mes_actual.columns = mes_actual.columns.str.strip().str.lower()

    pd.set_option('display.max_columns',None)
    pd.set_option('display.width', None)
    pd.set_option('display.max_colwidth', None)
    '''print(mes_actual.head())
    print(resumen.head())'''

    # Crear columnas de tipo de actividad
    mes_actual["horas_mantenimiento"] = mes_actual["cálculo horas"].where(mes_actual["activity"] == "MNT", 0)
    mes_actual["horas_cbk"] = mes_actual["cálculo horas"].where(mes_actual["activity"].isin(["CBK"]), 0)
    mes_actual["horas_extensivo"]=0


    # Agrupar por ID
    resumen_mes = mes_actual.groupby("id").agg({
        "horas_mantenimiento": "sum",
        "horas_cbk": "sum",
        "horas_extensivo": "sum",
    }).reset_index()
    
    # ---------- 3. Unir los resúmenes ----------
    final = pd.merge(resumen, resumen_mes, on="id", how="outer", suffixes=("_12meses", "_mes"))
    final = final.fillna(0)
    
    # Sumar totales
    final["horas_mantenimiento"] = final["horas_mantenimiento_12meses"] + final["horas_mantenimiento_mes"]
    final["horas_extensivo"] = final["horas_extensivo_12meses"] + final["horas_extensivo_mes"]
    final["horas_cbk"] = final["horas_cbk_12meses"] + final["horas_cbk_mes"]

    config.variable = 1
    # ---------- 4. Leer niveles de certificación ----------
    niveles=tratar_certificaciones(fichero2,opcion_seleccionada)

    '''print(niveles.head())'''
    niveles.rename(columns={"Personal ID":"Id"}, inplace=True)
    niveles.rename(columns={"Certification":"level"}, inplace=True)
    niveles.columns = niveles.columns.str.strip().str.lower()
    '''print(niveles.head())'''

    nivel_filtrado=niveles[["id","level"]]
    # Unir niveles
    final = pd.merge(final, nivel_filtrado, on="id", how="left")
    final["level"] = final["level"].fillna(0)
    final['level']=final['level'].astype(int)
    # ---------- 5. Aplicar reglas de certificación ----------
    def verificar_nivel(row):
        if (row["horas_cbk"] > 5 || row["horas_cbk_mes"]>1) and row["level"] < 2:
            return "CBK SIN L2"
        elif row["horas_extensivo"] > 5 and row["level"] < 1:
            return "EXTENSIVO SIN L1"
        elif row["horas_mantenimiento"] > 5 and row["level"] < 0:
            return "MNT SIN L0"
        else:
            return "OK"
    
    final["estado_certificacion"] = final.apply(verificar_nivel, axis=1)

    #Datos del ficheor de personal

    superPa=pd.read_excel(fichero3)
    superPa.rename(columns={"User/Employee ID":"id"}, inplace=True)
    superPa.columns=superPa.columns.str.strip().str.lower()
    superPafiltrado=superPa[["id","nombre completo","job title","job name","manager user sys id","supervisor","do","dr (dirección regional)","sucursal"]]

    final_completo=pd.merge(final,superPafiltrado,on="id", how="left")

    print("Fichero con incumplimientos de certificación creado exitosamennte, seleccione nombre y ubicación ")

    if variable==1:
        return final_completo
        break()
    
    # ---------- 6. Guardar con formato condicional en Excel ----------
    ruta_salida = seleccionar_ruta()
    if not ruta_salida:
        print("No se seleccionó ruta de salida")
        exit()
    final_completo.to_excel(ruta_salida, index=False)
    
    # Abrir con openpyxl y aplicar formato
    wb = load_workbook(ruta_salida)
    ws = wb.active
    
    # Buscar la columna de estado_certificacion
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "estado_certificacion":
            col_idx = col[0].column_letter
            break
    
    # Regla: si contiene "sin", poner en rojo
    formula = f'ISNUMBER(SEARCH("sin", ${col_idx}2))'
    red_font = Font(color="9C0006")
    rule = FormulaRule(formula=[formula], font=red_font)
    
    # Aplicar formato desde fila 2 hasta el final
    ws.conditional_formatting.add(f"{col_idx}2:{col_idx}{ws.max_row}", rule)
    
    # Guardar archivo final
    wb.save(ruta_salida)
    
    print("Archivo guardado correctamente")
    
