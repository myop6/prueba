import pandas as pd
import config

from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule

from seleccionar_ruta import seleccionar_ruta
from tratarHoras import tratarHoras
from tratar_certificaciones import tratar_certificaciones
from tratarHorasCompleto import tratarHorasCompleto
from tratarHorasEscaleras import tratarHorasEscaleras

def incumplimientosCompleto(horasCierreAsc,horasMesAsc,horasEsc,certificacion,personal,opcion_seleccionada):
    config.variable1=1
    ascensores=tratarHorasCompleto(horasCierreAsc,horasMesAsc,certificacion,personal,opcion_seleccionada)
    config.variable1=2
    escaleras=tratarHorasEscaleras(horasEsc,certificacion,personal,opcion_seleccionada)

    print("\nFichero con las desviaciones en las certificaciones creado correctamente. Seleccione nombre para el guardado\n")
    ruta_salida = seleccionar_ruta()

    if ruta_salida:
        with pd.ExcelWriter(ruta_salida) as writer:
            ascensores.to_excel(writer, sheet_name='Ascensores', index=False)
            escaleras.to_excel(writer, sheet_name='Escaleras', index=False)

    else:
        print('Guardado cancelado\n')



    # Abrir con openpyxl y aplicar formato
    wb = load_workbook(ruta_salida,)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Buscar la columna de estado_certificacion
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "estado_certificacion":
                col_idx = col[0].column_letter
                break
        # Regla: si contiene "sin", poner en rojo
        formula = f'ISNUMBER(SEARCH("sin", ${col_idx}2))'
        red_font = Font(color="9C0006")
        rule = FormulaRule(formula=[formula], font=red_font)
        # Aplicar formato desde fila 2 hasta el final
        ws.conditional_formatting.add(f"{col_idx}2:{col_idx}{ws.max_row}", rule)
        wb.save(ruta_salida)
    # Guardar archivo final
    wb.save(ruta_salida)


    print("Archivo guardado correctamente")